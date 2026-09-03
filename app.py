import io
import os
import threading
import time
from calendar import monthrange
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pytz import timezone

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import requests
from flask import Flask, render_template_string, jsonify, make_response, send_file
import json
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill

# ------------------- CONFIGURACIÓN -------------------
GCP_CREDENTIALS_JSON = os.getenv("GCP_CREDENTIALS_JSON")
SHEET_NAME = "verificacion_fechas"
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

# ✅ URL del propio servicio para el auto-ping (Cloud Run la inyecta automáticamente)
SERVICE_URL = os.getenv("SERVICE_URL", "")

# Segundos entre ciclos de verificación. Cada ciclo revisa UN bloque (ver
# BLOQUES más abajo), no las 17 bases enteras -- así nunca se manda una
# ráfaga grande de pings al servidor de la SBS de una sola vez.
INTERVALO_CICLO_SEGUNDOS = 60

# ------------------- LAS 17 BASES DEL PROYECTO SBS MULTIENTIDAD -------------------
# Mismos códigos que usan los pipelines de procesamiento (procesar_<base>.py).
# Algunas bases no incluyen EDPYMES porque esa familia no está autorizada a
# reportar ese tipo de dato (ej. Depósitos, Clientes de Ahorro, Ratio de
# Liquidez).
BASES = {
    "COLOCACIONES": {"BANCOS": "B-2334", "FINANCIERAS": "B-3220", "CMACS": "C-1228", "CRACS": "C-2228", "EDPYMES": "C-4223"},
    "DEPOSITOS": {"BANCOS": "B-2372", "FINANCIERAS": "B-3231", "CMACS": "C-1245", "CRACS": "C-2250"},
    "PERSONAL": {"BANCOS": "B-2305", "FINANCIERAS": "B-3202", "CMACS": "C-1202", "CRACS": "C-2202", "EDPYMES": "C-4206"},
    "CASTIGOS": {"BANCOS": "B-2369", "FINANCIERAS": "B-3234", "CMACS": "C-1253", "CRACS": "C-2258", "EDPYMES": "C-4242"},
    "CLIENTES_CREDITO": {"BANCOS": "B-230803", "FINANCIERAS": "B-3218", "CMACS": "C-1231", "CRACS": "C-2231", "EDPYMES": "C-4226"},
    "CLIENTES_AHORRO": {"BANCOS": "B-2373", "FINANCIERAS": "B-3232", "CMACS": "C-1250", "CRACS": "C-2255"},
    "CATEGORIA_RIESGO": {"BANCOS": "B-2309", "FINANCIERAS": "B-3205", "CMACS": "C-120201", "CRACS": "C-220201", "EDPYMES": "C-4201"},
    "PATRIMONIO_EFECTIVO": {"BANCOS": "B-2370", "FINANCIERAS": "B-3252", "CMACS": "C-1257", "CRACS": "C-2262", "EDPYMES": "C-4246"},
    "RCG": {"BANCOS": "B-2402", "FINANCIERAS": "B-3302", "CMACS": "C-1252", "CRACS": "C-2257", "EDPYMES": "C-4241"},
    "ESTRUCTURA_GASTO": {"BANCOS": "B-2390", "FINANCIERAS": "B-3253", "CMACS": "C-1239", "CRACS": "C-2244", "EDPYMES": "C-4233"},
    "INGRESOS_FINANCIEROS": {"BANCOS": "B-2347", "FINANCIERAS": "B-3224", "CMACS": "C-1220", "CRACS": "C-2220", "EDPYMES": "C-4215"},
    "RATIO_LIQUIDEZ": {"BANCOS": "B-2340", "FINANCIERAS": "B-3250", "CMACS": "C-1244", "CRACS": "C-2249"},
    "OFICINAS": {"BANCOS": "B-2303", "FINANCIERAS": "B-3201", "CMACS": "C-1201", "CRACS": "C-2201", "EDPYMES": "C-4205"},
    "CREDITOS_DEPOSITOS_ZONA": {"BANCOS": "B-2358", "FINANCIERAS": "B-3241", "CMACS": "C-1234", "CRACS": "C-2234", "EDPYMES": "C-4228"},
    "INDICADORES": {"BANCOS": "B-2401", "FINANCIERAS": "B-3301", "CMACS": "C-1301", "CRACS": "C-2301", "EDPYMES": "C-4301"},
    "GASTOS_ADMINISTRATIVOS": {"BANCOS": "B-2348", "FINANCIERAS": "B-3225", "CMACS": "C-1221", "CRACS": "C-2221", "EDPYMES": "C-4216"},
    "EEFF": {"BANCOS": "B-2201", "FINANCIERAS": "B-3101", "CMACS": "C-1101", "CRACS": "C-2101", "EDPYMES": "C-4103"},
}

# Nombres bonitos para mostrar en la interfaz
NOMBRES_BASE = {
    "COLOCACIONES": "Colocaciones", "DEPOSITOS": "Depósitos", "PERSONAL": "Personal",
    "CASTIGOS": "Castigos", "CLIENTES_CREDITO": "Clientes de Crédito",
    "CLIENTES_AHORRO": "Clientes de Ahorro", "CATEGORIA_RIESGO": "Categoría de Riesgo",
    "PATRIMONIO_EFECTIVO": "Patrimonio Efectivo", "RCG": "RCG",
    "ESTRUCTURA_GASTO": "Estructura de Gasto", "INGRESOS_FINANCIEROS": "Ingresos Financieros",
    "RATIO_LIQUIDEZ": "Ratio de Liquidez", "OFICINAS": "Oficinas por Zona Geográfica",
    "CREDITOS_DEPOSITOS_ZONA": "Créditos y Depósitos por Zona", "INDICADORES": "Indicadores",
    "GASTOS_ADMINISTRATIVOS": "Gastos Administrativos", "EEFF": "EEFF",
}
NOMBRES_FAMILIA = {
    "BANCOS": "Bancos", "FINANCIERAS": "Financieras", "CMACS": "CMACs",
    "CRACS": "CRACs", "EDPYMES": "EC",
}

# Bloques de revisión rotativa: cada ciclo solo procesa UN bloque, para no
# mandar una ráfaga de ~82 pings de golpe al servidor de la SBS.
BLOQUES = [
    ["COLOCACIONES", "DEPOSITOS", "CLIENTES_CREDITO", "CLIENTES_AHORRO"],
    ["PERSONAL", "CASTIGOS", "CATEGORIA_RIESGO", "PATRIMONIO_EFECTIVO"],
    ["RCG", "ESTRUCTURA_GASTO", "INGRESOS_FINANCIEROS", "RATIO_LIQUIDEZ"],
    ["OFICINAS", "CREDITOS_DEPOSITOS_ZONA", "INDICADORES", "GASTOS_ADMINISTRATIVOS", "EEFF"],
]

MESES = {
    1: ('Enero', 'en'), 2: ('Febrero', 'fe'), 3: ('Marzo', 'ma'), 4: ('Abril', 'ab'),
    5: ('Mayo', 'my'), 6: ('Junio', 'jn'), 7: ('Julio', 'jl'), 8: ('Agosto', 'ag'),
    9: ('Setiembre', 'se'), 10: ('Octubre', 'oc'), 11: ('Noviembre', 'no'), 12: ('Diciembre', 'di')
}
BOOTSTRAP_MESES_ATRAS = 8  # cuántos meses hacia atrás buscar al sembrar una fila nueva

# ------------------- ESTADO EN MEMORIA -------------------
# fechas_completas: {(base, familia): "dd/mm/YYYY" o None}
# filas_sheet: {(base, familia): número de fila en el sheet (1-indexed)}
fechas_completas = {}
filas_sheet = {}
estado_agrupado = {}
ultimo_envio = "Nunca"
bloque_actual = 0
_lock = threading.Lock()

HEADERS = ["ENTIDAD", "BASE", "FAMILIA", "FECHA", "ULTIMA_VERIFICACION"]


# ------------------- GOOGLE SHEETS -------------------
def conectar_google_sheet():
    if not GCP_CREDENTIALS_JSON:
        raise RuntimeError("⚠️ Variable GCP_CREDENTIALS_JSON no configurada")
    creds_dict = json.loads(GCP_CREDENTIALS_JSON)
    tmp_name = None
    try:
        with tempfile.NamedTemporaryFile(mode="w+", suffix=".json", delete=False) as tmp:
            json.dump(creds_dict, tmp)
            tmp_name = tmp.name
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(tmp_name, scope)
        client = gspread.authorize(creds)
        return client.open(SHEET_NAME).sheet1
    finally:
        if tmp_name and os.path.exists(tmp_name):
            os.unlink(tmp_name)


def _todas_las_combinaciones():
    """Genera (base, familia, codigo) para las 17 bases -- el orden define
    en qué fila cae cada una si hay que crearla nueva en el sheet."""
    for base, familias in BASES.items():
        for familia, codigo in familias.items():
            yield base, familia, codigo


def sincronizar_sheet(sheet):
    """
    Lee el sheet completo, arma fechas_completas/filas_sheet en memoria, y
    agrega (append) cualquier fila (base, familia) que todavía no exista --
    sembrando su fecha inicial con bootstrap_fecha(). Se corre una sola vez
    al arrancar el servicio.
    """
    global fechas_completas, filas_sheet

    valores = sheet.get_all_values()
    if not valores:
        sheet.update(values=[HEADERS], range_name="A1")
        valores = [HEADERS]

    headers_actuales = valores[0]
    if headers_actuales != HEADERS:
        # sheet viejo (13 entidades planas) -- no lo tocamos, solo agregamos
        # las filas nuevas después de lo que ya haya
        pass

    fechas_completas = {}
    filas_sheet = {}
    for i, fila in enumerate(valores[1:], start=2):
        if len(fila) < 5:
            continue
        base, familia, fecha = fila[1], fila[2], fila[3]
        if base and familia:
            fechas_completas[(base, familia)] = fecha or None
            filas_sheet[(base, familia)] = i

    # detectar combinaciones nuevas que falten y sembrarlas -- en paralelo,
    # porque son hasta 82 combinaciones x hasta 8 meses hacia atrás cada una
    # (podrían ser cientos de HEAD requests); en serie tardaría demasiado
    faltantes = [(base, familia, codigo) for base, familia, codigo in _todas_las_combinaciones()
                 if (base, familia) not in filas_sheet]

    resultados_bootstrap = {}
    if faltantes:
        print(f"🌱 Sembrando {len(faltantes)} filas nuevas (en paralelo)...")
        with ThreadPoolExecutor(max_workers=8) as executor:
            futuros = {executor.submit(bootstrap_fecha, codigo): (base, familia)
                       for base, familia, codigo in faltantes}
            for futuro in as_completed(futuros):
                base, familia = futuros[futuro]
                try:
                    resultados_bootstrap[(base, familia)] = futuro.result()
                except Exception as e:
                    print(f"⚠️ Bootstrap falló para {base}/{familia}: {e}")
                    resultados_bootstrap[(base, familia)] = None

    filas_nuevas = []
    siguiente_fila = len(valores) + 1
    for base, familia, codigo in faltantes:
        key = (base, familia)
        fecha_inicial = resultados_bootstrap.get(key)
        entidad_key = f"{base}__{familia}"
        filas_nuevas.append([entidad_key, base, familia, fecha_inicial or "", ""])
        fechas_completas[key] = fecha_inicial
        filas_sheet[key] = siguiente_fila
        siguiente_fila += 1

    if filas_nuevas:
        rango = f"A{len(valores) + 1}:E{len(valores) + len(filas_nuevas)}"
        sheet.update(values=filas_nuevas, range_name=rango)
        print(f"✅ {len(filas_nuevas)} filas nuevas agregadas al sheet")


def escribir_cambios_sheet(sheet, cambios: dict, timestamp: str):
    """
    cambios: {(base, familia): nueva_fecha_str}
    Escribe SOLO las filas que cambiaron en este ciclo, en un solo batch_update
    (una llamada a la API en vez de una por celda).
    """
    if not cambios:
        return
    updates = []
    for key, nueva_fecha in cambios.items():
        fila = filas_sheet.get(key)
        if not fila:
            continue
        updates.append({
            "range": f"D{fila}:E{fila}",
            "values": [[nueva_fecha, timestamp]],
        })
    if updates:
        sheet.batch_update(updates)


# ------------------- TELEGRAM -------------------
def enviar_telegram(mensaje):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("⚠️ No hay configuración de Telegram")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": mensaje}
    try:
        requests.post(url, json=payload, timeout=10)
        print(f"📲 Telegram enviado: {mensaje}")
    except Exception as e:
        print(f"❌ Error enviando Telegram: {e}")


# ------------------- LÓGICA DE FECHAS Y VERIFICACIÓN -------------------
def es_fecha_valida(fecha):
    if not fecha:
        return False
    try:
        datetime.strptime(fecha, "%d/%m/%Y")
        return True
    except Exception:
        return False


def obtener_mes_siguiente(fecha_str):
    try:
        fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
        anio, mes = fecha.year, fecha.month + 1
        if mes > 12:
            mes, anio = 1, anio + 1
        return anio, mes
    except Exception:
        return None, None


def verificar_archivo_codigo(anio, mes, codigo):
    mes_nombre, mes_abr = MESES[mes]
    url = f"https://intranet2.sbs.gob.pe/estadistica/financiera/{anio}/{mes_nombre}/{codigo}-{mes_abr}{anio}.xls"
    try:
        response = requests.head(url, timeout=5)
        return response.status_code == 200
    except Exception:
        return None


def bootstrap_fecha(codigo, meses_atras=BOOTSTRAP_MESES_ATRAS):
    """Para una fila nueva sin fecha conocida: prueba mes por mes hacia atrás
    desde hoy hasta encontrar el archivo más reciente que sí existe."""
    hoy = datetime.now(timezone('America/Lima'))
    anio, mes = hoy.year, hoy.month
    for _ in range(meses_atras):
        if verificar_archivo_codigo(anio, mes, codigo):
            dia_final = monthrange(anio, mes)[1]
            return f"{dia_final:02d}/{mes:02d}/{anio}"
        mes -= 1
        if mes < 1:
            mes, anio = 12, anio - 1
    return None


def formato_corto(fecha_str):
    if not es_fecha_valida(fecha_str):
        return "—"
    fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
    mes_nombre = MESES[fecha.month][0][:3]
    return f"{mes_nombre}{str(fecha.year)[-2:]}"


def construir_estado_agrupado():
    """A partir de fechas_completas (todas las bases), arma la vista por
    grupo: fecha máxima del grupo (la familia más adelantada) y si cada
    familia está al día respecto a ESA fecha máxima -- no respecto a un
    calendario fijo. Ej.: si RCG-Bancos ya salió Jul26 pero el resto de RCG
    sigue en Jun26, el grupo RCG muestra "Jul26" como máxima, Bancos en
    verde, el resto en rojo."""
    estado = {}
    for base, familias in BASES.items():
        fechas_dt = {}
        for familia in familias:
            fecha_str = fechas_completas.get((base, familia))
            if es_fecha_valida(fecha_str):
                fechas_dt[familia] = datetime.strptime(fecha_str, "%d/%m/%Y")

        fecha_max_dt = max(fechas_dt.values()) if fechas_dt else None
        familias_estado = {}
        for familia in familias:
            fecha_str = fechas_completas.get((base, familia))
            fecha_dt = fechas_dt.get(familia)
            al_dia = bool(fecha_dt and fecha_max_dt and fecha_dt == fecha_max_dt)
            familias_estado[familia] = {
                "nombre": NOMBRES_FAMILIA[familia],
                "fecha": formato_corto(fecha_str),
                "al_dia": al_dia,
            }
        estado[base] = {
            "nombre": NOMBRES_BASE[base],
            "fecha_max": formato_corto(fecha_max_dt.strftime("%d/%m/%Y")) if fecha_max_dt else "—",
            "familias": familias_estado,
        }
    return estado


def revisar_bloque(sheet, bases_del_bloque):
    """Revisa solo las bases de un bloque -- un HEAD request por cada
    combinación (base, familia) de ese bloque, no de las 17 bases enteras."""
    timestamp_actual = datetime.now(timezone('America/Lima')).strftime('%Y-%m-%d %H:%M:%S %Z')
    cambios = {}
    nuevos_archivos = []

    for base in bases_del_bloque:
        for familia, codigo in BASES[base].items():
            key = (base, familia)
            fecha_actual = fechas_completas.get(key)

            if not es_fecha_valida(fecha_actual):
                # nunca se pudo sembrar (la SBS no tenía nada en la ventana de
                # bootstrap) -- se reintenta el bootstrap en este ciclo
                fecha_sembrada = bootstrap_fecha(codigo)
                if fecha_sembrada:
                    cambios[key] = fecha_sembrada
                    fechas_completas[key] = fecha_sembrada
                continue

            anio, mes = obtener_mes_siguiente(fecha_actual)
            if not anio:
                continue
            existe = verificar_archivo_codigo(anio, mes, codigo)
            if existe:
                dia_final = monthrange(anio, mes)[1]
                nueva_fecha = f"{dia_final:02d}/{mes:02d}/{anio}"
                cambios[key] = nueva_fecha
                fechas_completas[key] = nueva_fecha
                nuevos_archivos.append(f"{NOMBRES_BASE[base]} · {NOMBRES_FAMILIA[familia]} → {formato_corto(nueva_fecha)}")

    if cambios:
        escribir_cambios_sheet(sheet, cambios, timestamp_actual)

    if nuevos_archivos:
        mensaje = "📁 Nuevos archivos SBS detectados:\n" + "\n".join(f"• {linea}" for linea in nuevos_archivos)
        enviar_telegram(mensaje)

    return timestamp_actual


# ------------------- CICLOS EN SEGUNDO PLANO -------------------
def ciclo_verificacion():
    global estado_agrupado, ultimo_envio, bloque_actual

    time.sleep(5)
    sheet = None
    while sheet is None:
        try:
            sheet = conectar_google_sheet()
            sincronizar_sheet(sheet)
            with _lock:
                estado_agrupado = construir_estado_agrupado()
        except Exception as e:
            print(f"❌ Error en la sincronización inicial, reintentando en 30s: {e}")
            sheet = None
            time.sleep(30)

    while True:
        try:
            bloque = BLOQUES[bloque_actual % len(BLOQUES)]
            print(f"🔍 Revisando bloque {bloque_actual % len(BLOQUES) + 1}/{len(BLOQUES)}: {bloque}")
            ts = revisar_bloque(sheet, bloque)
            with _lock:
                estado_agrupado = construir_estado_agrupado()
                ultimo_envio = ts
            bloque_actual += 1
            print(f"✅ Bloque revisado: {ts}")
        except Exception as e:
            print(f"❌ Error en verificación: {e}")
        time.sleep(INTERVALO_CICLO_SEGUNDOS)


def ciclo_keepalive():
    """Ping a sí mismo cada 4 minutos para que Cloud Run no escale a cero la
    instancia y mate el hilo de verificación."""
    time.sleep(30)
    while True:
        try:
            if SERVICE_URL:
                requests.get(f"{SERVICE_URL}/healthz", timeout=10)
                print(f"🏓 Keepalive ping OK: {datetime.now(timezone('America/Lima')).strftime('%H:%M:%S')}")
            else:
                print("⚠️ SERVICE_URL no configurada, keepalive desactivado")
        except Exception as e:
            print(f"⚠️ Keepalive falló: {e}")
        time.sleep(4 * 60)


# ------------------- SERVIDOR WEB -------------------
app = Flask(__name__)


@app.route("/healthz")
def healthz():
    return "ok", 200


@app.route("/data")
def data():
    with _lock:
        payload = {"ultimo_envio": ultimo_envio, "bases": estado_agrupado}
    resp = make_response(jsonify(payload))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/excel")
def excel():
    with _lock:
        estado = dict(estado_agrupado)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado SBS"
    ws.append(["Base", "Fecha máxima del grupo", "Familia", "Fecha cargada", "Al día"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    fila = 2
    for base, info in estado.items():
        for familia, f_info in info["familias"].items():
            ws.append([info["nombre"], info["fecha_max"], f_info["nombre"], f_info["fecha"],
                       "Sí" if f_info["al_dia"] else "No"])
            ws.cell(row=fila, column=5).fill = verde if f_info["al_dia"] else rojo
            fila += 1

    for col in ws.columns:
        largo = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = largo + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"estado_sbs_{datetime.now(timezone('America/Lima')).strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=nombre_archivo,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Verificación SBS</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background: #f4f6f9; }
        .card-base { border-radius: 14px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
        .card-base .card-header { border-radius: 14px 14px 0 0; font-weight: 600; }
        .badge-familia { font-size: 0.85rem; padding: 0.5em 0.8em; }
        .fecha-max { font-weight: 700; }
    </style>
</head>
<body>
<div class="container py-4">
    <div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
        <h1 class="mb-0">📊 Verificación de Archivos SBS</h1>
        <a href="/excel" class="btn btn-primary">⬇️ Descargar Excel</a>
    </div>
    <p class="text-muted">Última verificación: <strong id="ultimo_envio">{{ ultimo_envio }}</strong></p>

    <div class="row row-cols-1 row-cols-md-2 row-cols-lg-3 g-3" id="tarjetas_bases"></div>
</div>

<script>
    function pintarTarjetas(bases) {
        const cont = document.getElementById("tarjetas_bases");
        cont.innerHTML = "";
        for (const base in bases) {
            const info = bases[base];
            let badges = "";
            for (const fam in info.familias) {
                const f = info.familias[fam];
                const color = f.al_dia ? "bg-success" : "bg-danger";
                badges += `<span class="badge ${color} badge-familia me-1 mb-1">${f.nombre}: ${f.fecha}</span>`;
            }
            const col = document.createElement("div");
            col.className = "col";
            col.innerHTML = `
                <div class="card card-base h-100">
                    <div class="card-header bg-white d-flex justify-content-between align-items-center">
                        <span>${info.nombre}</span>
                        <span class="fecha-max text-primary">${info.fecha_max}</span>
                    </div>
                    <div class="card-body">${badges}</div>
                </div>`;
            cont.appendChild(col);
        }
    }

    function actualizarDatos() {
        fetch("/data")
            .then(r => r.json())
            .then(data => {
                document.getElementById("ultimo_envio").innerText = data.ultimo_envio;
                pintarTarjetas(data.bases);
            })
            .catch(err => console.error("Error actualizando datos:", err));
    }
    setInterval(actualizarDatos, 10000);
    actualizarDatos();
</script>
</body>
</html>
"""


@app.route("/")
def home():
    with _lock:
        ts = ultimo_envio
    return render_template_string(TEMPLATE, ultimo_envio=ts)


# ✅ Arranque de hilos — sin código síncrono pesado aquí
threading.Thread(target=ciclo_verificacion, daemon=True).start()
threading.Thread(target=ciclo_keepalive, daemon=True).start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
